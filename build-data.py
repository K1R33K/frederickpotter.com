#!/usr/bin/env python3
"""
Reads 'Geeks Ltd - My Case Studies.xlsx' and generates 'js/data.js'.
Run this whenever the spreadsheet changes:
    python3 portfolio/build-data.py
"""

import json
import os
import re
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "..", "Geeks Ltd - My Case Studies.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "js", "data.js")

# ── Sector assignments (case study number → sector key) ──────────────────────
SECTOR_MAP = {
    # Property & Real Estate
    1: "proptech", 2: "proptech", 3: "proptech", 42: "proptech",
    # Climate, Energy & Sustainability
    4: "climate-energy", 13: "climate-energy", 14: "climate-energy", 85: "climate-energy",
    # Government & Public Sector
    5: "govpublic", 35: "govpublic", 37: "govpublic", 50: "govpublic",
    54: "govpublic", 67: "govpublic", 81: "govpublic", 101: "govpublic",
    106: "govpublic",
    # Financial Services & FinTech
    20: "fintech", 27: "fintech", 36: "fintech", 40: "fintech",
    43: "fintech", 45: "fintech", 62: "fintech", 65: "fintech",
    72: "fintech", 82: "fintech",
    # Logistics, Trade & Customs
    9: "logistics", 10: "logistics", 11: "logistics", 12: "logistics", 79: "logistics",
    105: "logistics",
    # Healthcare & Life Sciences
    23: "health", 29: "health", 47: "health", 59: "health",
    60: "health", 92: "health", 99: "health", 102: "health",
    107: "health", 108: "health",
    # Education & Training
    7: "education", 8: "education", 15: "education", 34: "education", 90: "education",
    # Legal & Professional Services
    32: "legal", 39: "legal", 55: "legal", 52: "legal",
    # Construction & Engineering
    18: "construction", 33: "construction", 41: "construction", 57: "construction",
    75: "construction", 80: "construction",
    # Media, Brand & Entertainment
    21: "media", 22: "media", 26: "media", 38: "media",
    44: "media", 49: "media", 51: "media", 63: "media", 100: "media",
    # Social Impact & Charities
    17: "social-impact", 68: "social-impact", 69: "social-impact",
    70: "social-impact", 71: "social-impact", 77: "social-impact", 88: "social-impact",
    # Recruitment, HR & Workforce
    24: "recruitment", 30: "recruitment", 31: "recruitment",
    61: "recruitment", 66: "recruitment", 86: "recruitment",
    95: "recruitment", 96: "recruitment",
    # Technology & AI
    6: "technology", 16: "technology", 19: "technology", 25: "technology",
    28: "technology", 46: "technology", 48: "technology", 53: "technology",
    56: "technology", 58: "technology", 64: "technology", 73: "technology",
    74: "technology", 76: "technology", 78: "technology", 83: "technology",
    84: "technology", 87: "technology", 89: "technology", 91: "technology",
    93: "technology", 94: "technology", 97: "technology", 98: "technology",
    103: "technology", 104: "technology",
}

SECTOR_LABELS = {
    "proptech":       {"name": "Property & Real Estate",     "icon": "building"},
    "climate-energy": {"name": "Climate, Energy & Sustainability", "icon": "leaf"},
    "govpublic":      {"name": "Government & Public Sector", "icon": "landmark"},
    "fintech":        {"name": "Financial Services",         "icon": "chart-line"},
    "logistics":      {"name": "Logistics, Trade & Customs", "icon": "truck"},
    "health":         {"name": "Healthcare & Life Sciences", "icon": "heart-pulse"},
    "education":      {"name": "Education & Training",       "icon": "graduation-cap"},
    "legal":          {"name": "Legal & Professional Services", "icon": "scale"},
    "construction":   {"name": "Construction & Engineering", "icon": "hard-hat"},
    "media":          {"name": "Media, Brand & Entertainment", "icon": "play"},
    "social-impact":  {"name": "Social Impact & Charities",  "icon": "hand-heart"},
    "recruitment":    {"name": "Recruitment, HR & Workforce", "icon": "users"},
    "technology":     {"name": "Technology & AI",             "icon": "cpu"},
}

# ── Engagement type (derived from relationship field) ─────────────────────────
ENGAGEMENT_LABELS = {
    "repeat":      {"name": "Repeat Client"},
    "partnership":  {"name": "Long-term Partnership"},
    "single":       {"name": "Single Engagement"},
}

def derive_engagement_type(relationship):
    """Derive engagement type from the relationship field text."""
    r = (relationship or "").lower()
    if r.startswith("repeat client"):
        return "repeat"
    if r.startswith("multi-phase") or r.startswith("ongoing") or r.startswith("connected entity"):
        return "partnership"
    return "single"

# ── Project type (case study number → list of type keys) ─────────────────────
# A project can have multiple types.
PROJECT_TYPE_MAP = {
    # AI & Machine Learning (57 projects total)
    1: ["ai-ml", "build"], 2: ["ai-ml", "build"], 3: ["ai-ml", "build"],
    4: ["ai-ml", "strategy"], 6: ["ai-ml", "build"],
    7: ["ai-ml", "strategy"], 8: ["ai-ml", "strategy"],
    9: ["ai-ml", "build"], 10: ["ai-ml", "build"],
    12: ["ai-ml", "build"], 13: ["ai-ml", "strategy"],
    14: ["ai-ml", "strategy"], 16: ["ai-ml", "build"],
    17: ["ai-ml", "build"], 19: ["ai-ml", "strategy"],
    20: ["ai-ml", "build"], 24: ["ai-ml", "strategy"],
    30: ["ai-ml", "build"], 36: ["ai-ml", "build"],
    37: ["ai-ml", "build"], 38: ["ai-ml", "build"],
    39: ["ai-ml", "build"], 41: ["ai-ml", "build"],
    57: ["ai-ml", "build"], 59: ["ai-ml", "build"],
    67: ["ai-ml", "build"], 72: ["ai-ml", "build", "data"],
    73: ["ai-ml", "build"], 74: ["ai-ml", "build", "data"],
    75: ["ai-ml"], 76: ["ai-ml", "build"],
    77: ["ai-ml", "build"], 78: ["ai-ml", "build"],
    79: ["ai-ml"], 80: ["ai-ml", "build"],
    81: ["ai-ml", "build"], 83: ["ai-ml", "build"],
    84: ["ai-ml", "build", "data"], 85: ["ai-ml", "build", "data"],
    88: ["ai-ml", "build"], 90: ["ai-ml", "build"],
    91: ["ai-ml", "strategy"], 92: ["ai-ml", "build"],
    93: ["ai-ml", "build", "data"], 94: ["ai-ml", "build"],
    95: ["ai-ml", "build"], 96: ["ai-ml", "build"],
    99: ["ai-ml", "build"], 100: ["ai-ml", "build", "data"],
    103: ["ai-ml", "data"], 104: ["ai-ml", "data"],
    105: ["ai-ml", "build", "data"], 106: ["ai-ml", "build"],
    107: ["ai-ml", "build", "data"],
    # AI & Machine Learning + Data focus
    35: ["ai-ml", "build", "data"], 43: ["ai-ml", "build", "data"],
    97: ["ai-ml", "build", "data"],
    # Strategy & Advisory (pure strategy, no AI)
    21: ["strategy"], 22: ["strategy"], 25: ["strategy"],
    31: ["strategy"], 32: ["strategy"], 33: ["strategy"],
    56: ["strategy"], 69: ["strategy"],
    # Product Build (no AI, no significant data focus)
    5: ["build"], 11: ["build"], 15: ["build"], 18: ["build"],
    23: ["build"], 26: ["build"], 27: ["build"], 28: ["build"],
    29: ["build"], 34: ["build"], 40: ["build"], 42: ["build"],
    44: ["build"], 46: ["build"], 47: ["build"], 48: ["build"],
    50: ["build"], 51: ["build"], 52: ["build"], 53: ["build"],
    55: ["build"], 58: ["build"], 61: ["build"], 62: ["build"],
    63: ["build"], 64: ["build"], 66: ["build"], 68: ["build"],
    70: ["build"], 71: ["build"], 82: ["build"], 86: ["build"],
    87: ["build"], 98: ["build"], 101: ["build"], 102: ["build"],
    # Product Build + Data focus (no AI)
    45: ["build", "data"], 49: ["build", "data"], 54: ["build", "data"],
    60: ["build", "data"], 65: ["build", "data"], 89: ["build", "data"],
}

PROJECT_TYPE_LABELS = {
    "ai-ml":     {"name": "AI & Machine Learning"},
    "strategy":  {"name": "Strategy & Advisory"},
    "build":     {"name": "Product Build"},
    "data":      {"name": "Data & Analytics"},
}

# ── Company size / type (case study number → size key) ────────────────────────
COMPANY_SIZE_MAP = {
    # Startup / Scale-up
    1: "startup", 2: "startup", 3: "startup",       # Search Acumen (PropTech, founded 2013)
    4: "startup",                                     # Vaulted Deep (climate tech)
    47: "startup",                                    # Harley Street Connect
    49: "startup",                                    # Histropedia (founded 2012)
    58: "startup",                                    # Locpin (father-and-son founders)
    62: "startup",                                    # Mortgagez
    66: "startup",                                    # WorkGaps
    72: "startup",                                    # Unetix (AI analytics)
    80: "startup",                                    # IDM (construction marketplace)
    81: "startup",                                    # Incumbency AI
    82: "startup",                                    # Boodle (cost-sharing app)
    83: "startup",                                    # FYIO (digital filing)
    85: "startup",                                    # Wattsight (AI startup)
    88: "startup",                                    # The Fundraising Accountant
    93: "startup",                                    # Scoutd AI
    95: "startup",                                    # Prime Pro (recruitment CRM)
    96: "startup",                                    # Prime Pay (payroll platform)
    100: "startup",                                   # Birchbox
    102: "startup",                                   # eMed / Babylon Health
    # SME
    5: "sme", 6: "sme",                              # VPP, Prestige Purchasing
    7: "sme", 8: "sme",                              # Lord Wandsworth College
    10: "sme", 11: "sme", 12: "sme",                 # ChannelPorts
    13: "sme", 14: "sme",                            # Ignition Group
    15: "sme",                                        # GDAK (cybersecurity)
    16: "sme",                                        # Champion Timber
    18: "sme",                                        # Houston Cox
    20: "sme",                                        # Genesis Capital
    21: "sme", 22: "sme",                            # TSL (broadcast)
    25: "sme",                                        # EPIC Global Solutions
    26: "sme",                                        # Big Bad Wolf (agency)
    28: "sme",                                        # Fixatex
    29: "sme",                                        # Tristel
    30: "sme",                                        # 247 Time
    32: "sme",                                        # Morr & Co
    33: "sme",                                        # Paragon Building (130 people)
    36: "sme",                                        # Eazy Collect
    39: "sme",                                        # SSW
    40: "sme",                                        # Lucid Issuer Services
    41: "sme",                                        # CMS Desk
    42: "sme",                                        # Arthur Beeston Law (11-50)
    43: "sme",                                        # Brand Finance
    48: "sme",                                        # CET (InsurTech)
    52: "sme",                                        # Lawmens (family business)
    53: "sme",                                        # Commtel / Telguard
    55: "sme",                                        # PracticeEvolve
    56: "sme",                                        # Niche Care
    59: "sme",                                        # CPS (clinical lab)
    65: "sme",                                        # SALAMANCA Group
    73: "sme",                                        # Fixatex (later project)
    75: "sme",                                        # RAD Fire Sprinklers
    79: "sme",                                        # FFCL
    86: "sme",                                        # UCAN
    87: "sme",                                        # THSP (30+ years)
    90: "sme",                                        # English School of Kuwait
    98: "sme",                                        # Sleeping Giant Media
    99: "sme",                                        # DLS Health
    106: "sme",                                       # RC Fornax (Defence consulting)
    107: "sme",                                       # ila (Skincare AI)
    # Enterprise / Corporate
    9: "enterprise",                                  # va-Q-tec (listed, FT Tech Champion)
    19: "enterprise",                                 # Birkin Group (1000-5000 employees)
    23: "enterprise", 24: "enterprise",               # Reed group
    27: "enterprise",                                 # Guy Carpenter (Marsh & McLennan)
    31: "enterprise",                                 # People2.0 (global)
    38: "enterprise",                                 # BBC Mundo / BBC World
    44: "enterprise",                                 # easyJet (55M+ passengers)
    45: "enterprise",                                 # Euromoney (100+ magazines)
    46: "enterprise",                                 # Payzone (Post Office group)
    51: "enterprise",                                 # HMV (120+ stores)
    57: "enterprise",                                 # AESSEAL (4th largest globally)
    60: "enterprise",                                 # LSG (UK's largest provider)
    61: "enterprise",                                 # MITIE (FTSE 250)
    63: "enterprise",                                 # Rexam PLC (global packaging)
    78: "enterprise",                                 # Westlands (defence)
    84: "enterprise",                                 # Aker BP (major oil & gas)
    89: "enterprise",                                 # Dyson Farming
    91: "enterprise",                                 # Weybourne ($17B family office)
    92: "enterprise",                                 # PAM (largest privately owned)
    94: "enterprise",                                 # Porsche
    105: "enterprise",                                # Getlink (Channel Tunnel operator)
    # Government & Public Sector
    34: "public-sector",                              # UAL
    35: "public-sector",                              # City of Westminster
    37: "public-sector",                              # Council of Europe
    50: "public-sector",                              # Visit Britain
    54: "public-sector",                              # Houses of Parliament
    64: "public-sector",                              # RCGP (42,000 members)
    67: "public-sector",                              # ARB (established by Parliament)
    101: "public-sector",                             # NHS COVID-19 Vaccination
    # Non-profit / Charity
    17: "charity",                                    # Big Life Adventure (social enterprise)
    68: "charity",                                    # Coram
    69: "charity",                                    # Sands
    70: "charity",                                    # Safer Places
    71: "charity",                                    # Shooting Star Children's Hospices
    77: "charity",                                    # Fort Bend Women's Center
    # Freelance / Personal
    74: "freelance",                                  # Legion (personal project)
    76: "freelance",                                  # Internal tooling
    97: "freelance",                                  # Range Hunter (freelance)
    103: "freelance",                                 # Synthetic Financial Data (GAN)
    104: "freelance",                                 # Synthetic Financial Data (Diffusion)
}

COMPANY_SIZE_LABELS = {
    "startup":       {"name": "Startup / Scale-up"},
    "sme":           {"name": "SME"},
    "enterprise":    {"name": "Enterprise / Corporate"},
    "public-sector": {"name": "Government & Public Sector"},
    "charity":       {"name": "Non-profit / Charity"},
    "freelance":     {"name": "Freelance / Personal"},
}

# ── Spotlight projects ────────────────────────────────────────────────────────
SPOTLIGHT_NUMS = {2, 10, 27, 37, 44, 46, 54, 63, 64, 74, 84, 89, 91, 94, 101}


def parse_client_project(raw):
    """Split 'Client\\nProject Name' into (client, project)."""
    if "\n" in raw:
        parts = raw.split("\n", 1)
        return parts[0].strip(), parts[1].strip()
    # No newline; try splitting on " - "
    if " - " in raw:
        parts = raw.split(" - ", 1)
        return parts[0].strip(), parts[1].strip()
    return raw.strip(), ""


def parse_skills(raw):
    """Split comma-separated skills into a clean list."""
    if not raw:
        return []
    return [s.strip() for s in raw.split(",") if s.strip()]


def extract_metrics(result, action):
    """Pull headline stats and quotes from result/action text."""
    metrics = {}

    # Try to find a quoted testimonial
    quote_patterns = [
        r"[''\"](.*?)[''\"]\s*$",           # trailing quote
        r":\s*[''\"](.*?)[''\"]\s*$",        # colon then quote
        r"[''\"](.*?)[''\"]\s*$",            # unicode quotes
    ]
    text = result or ""
    for pattern in quote_patterns:
        match = re.search(pattern, text)
        if match and len(match.group(1)) > 15:
            metrics["quote"] = match.group(1)
            break

    # Fall back to action text for quotes
    if "quote" not in metrics and action:
        for pattern in quote_patterns:
            match = re.search(pattern, action)
            if match and len(match.group(1)) > 15:
                metrics["quote"] = match.group(1)
                break

    return metrics


def read_spreadsheet():
    """Read the Excel file and return structured case study data."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["My Case Studies"]

    case_studies = []
    for row_idx in range(5, ws.max_row + 1):
        num_val = ws.cell(row=row_idx, column=1).value
        if num_val is None or not str(num_val).strip().isdigit():
            continue

        num = int(str(num_val).strip())
        raw_client_project = str(ws.cell(row=row_idx, column=2).value or "")
        client, project = parse_client_project(raw_client_project)

        cs = {
            "num": num,
            "client": client,
            "project": project,
            "about": str(ws.cell(row=row_idx, column=3).value or "").strip(),
            "challenge": str(ws.cell(row=row_idx, column=4).value or "").strip(),
            "action": str(ws.cell(row=row_idx, column=5).value or "").strip(),
            "skills": parse_skills(str(ws.cell(row=row_idx, column=6).value or "")),
            "result": str(ws.cell(row=row_idx, column=7).value or "").strip(),
            "relationship": str(ws.cell(row=row_idx, column=8).value or "").strip(),
            "sector": SECTOR_MAP.get(num, "technology"),
            "spotlight": num in SPOTLIGHT_NUMS,
            "engagementType": derive_engagement_type(
                str(ws.cell(row=row_idx, column=8).value or "")
            ),
            "projectTypes": PROJECT_TYPE_MAP.get(num, ["build"]),
            "companySize": COMPANY_SIZE_MAP.get(num, "sme"),
        }

        if cs["spotlight"]:
            cs["metrics"] = extract_metrics(cs["result"], cs["action"])

        case_studies.append(cs)

    return case_studies


def generate_js(case_studies):
    """Write the data.js file."""
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    # Build sector counts
    sector_counts = {}
    for cs in case_studies:
        sector_counts[cs["sector"]] = sector_counts.get(cs["sector"], 0) + 1

    # Build skill frequency map
    skill_freq = {}
    for cs in case_studies:
        for skill in cs["skills"]:
            skill_freq[skill] = skill_freq.get(skill, 0) + 1

    # Sort skills by frequency descending
    sorted_skills = sorted(skill_freq.items(), key=lambda x: -x[1])

    # Build engagement type counts
    engagement_counts = {}
    for cs in case_studies:
        et = cs["engagementType"]
        engagement_counts[et] = engagement_counts.get(et, 0) + 1

    # Build project type counts
    project_type_counts = {}
    for cs in case_studies:
        for pt in cs["projectTypes"]:
            project_type_counts[pt] = project_type_counts.get(pt, 0) + 1

    # Build company size counts
    company_size_counts = {}
    for cs in case_studies:
        sz = cs["companySize"]
        company_size_counts[sz] = company_size_counts.get(sz, 0) + 1

    js_lines = [
        "// Auto-generated by build-data.py. Do not edit manually.",
        f"// Generated from: Geeks Ltd - My Case Studies.xlsx",
        f"// Total case studies: {len(case_studies)}",
        "",
        f"export const CASE_STUDIES = {json.dumps(case_studies, indent=2, ensure_ascii=False)};",
        "",
        f"export const SECTOR_LABELS = {json.dumps(SECTOR_LABELS, indent=2, ensure_ascii=False)};",
        "",
        f"export const SECTOR_COUNTS = {json.dumps(sector_counts, indent=2)};",
        "",
        f"export const SKILL_FREQUENCIES = {json.dumps(dict(sorted_skills), indent=2, ensure_ascii=False)};",
        "",
        f"export const ENGAGEMENT_LABELS = {json.dumps(ENGAGEMENT_LABELS, indent=2, ensure_ascii=False)};",
        "",
        f"export const ENGAGEMENT_COUNTS = {json.dumps(engagement_counts, indent=2)};",
        "",
        f"export const PROJECT_TYPE_LABELS = {json.dumps(PROJECT_TYPE_LABELS, indent=2, ensure_ascii=False)};",
        "",
        f"export const PROJECT_TYPE_COUNTS = {json.dumps(project_type_counts, indent=2)};",
        "",
        f"export const COMPANY_SIZE_LABELS = {json.dumps(COMPANY_SIZE_LABELS, indent=2, ensure_ascii=False)};",
        "",
        f"export const COMPANY_SIZE_COUNTS = {json.dumps(company_size_counts, indent=2)};",
        "",
        f"export const TOTAL_PROJECTS = {len(case_studies)};",
        f"export const TOTAL_SKILLS = {len(skill_freq)};",
        f"export const TOTAL_SECTORS = {len(SECTOR_LABELS)};",
        "",
    ]

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(js_lines))

    print(f"Generated {OUTPUT_PATH}")
    print(f"  {len(case_studies)} case studies")
    print(f"  {len(SECTOR_LABELS)} sectors")
    print(f"  {len(skill_freq)} unique skills")
    print(f"  {sum(1 for cs in case_studies if cs['spotlight'])} spotlight projects")

    # Warn about missing sector assignments
    missing = [cs["num"] for cs in case_studies if cs["num"] not in SECTOR_MAP]
    if missing:
        print(f"  WARNING: {len(missing)} case studies defaulted to 'technology': {missing}")


if __name__ == "__main__":
    case_studies = read_spreadsheet()
    generate_js(case_studies)
